from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone

from apps.campaigns.models import CampaignForm


BRAND_FILL = PatternFill("solid", fgColor="2B2623")
HEADER_FILL = PatternFill("solid", fgColor="D9B88C")
SUBTLE_FILL = PatternFill("solid", fgColor="F6EEE4")
WHITE_FONT = Font(color="FFF7F0", bold=True)
TITLE_FONT = Font(color="FFF7F0", bold=True, size=16)
HEADER_FONT = Font(color="2B2623", bold=True)
BODY_FONT = Font(color="2B2623")
THIN_BOTTOM = Border(bottom=Side(style="thin", color="D9B88C"))


def build_campaign_responses_workbook(form: CampaignForm) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Responses"
    sheet.sheet_view.showGridLines = False

    fields = list(form.fields.order_by("ordering", "id"))
    responses = list(form.responses.order_by("submitted_at"))
    headers = ["Submitted At", "Response ID"] + [field.label for field in fields]
    last_column = get_column_letter(len(headers))

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"] = "The Glow Mission"
    sheet["A1"].fill = BRAND_FILL
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].alignment = Alignment(horizontal="center")
    for column in range(2, len(headers) + 1):
        sheet.cell(row=1, column=column).fill = BRAND_FILL

    sheet["A2"] = f"{form.title} responses"
    sheet["A3"] = f"Exported: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')}"
    sheet["A4"] = f"Response count: {len(responses)}"
    for row in range(2, 5):
        sheet.cell(row=row, column=1).font = HEADER_FONT
        sheet.cell(row=row, column=1).fill = SUBTLE_FILL
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))

    header_row = 6
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BOTTOM

    for row_index, response in enumerate(responses, start=header_row + 1):
        submitted_at = timezone.localtime(response.submitted_at).replace(tzinfo=None)
        row_values = [submitted_at, response.id]
        for field in fields:
            row_values.append(format_response_value(response.response_data.get(field.key, "")))
        for column, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column == 1:
                cell.number_format = "yyyy-mm-dd hh:mm"

    last_row = max(header_row + 1, header_row + len(responses))
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
    sheet.freeze_panes = f"B{header_row + 1}"

    autosize_columns(sheet, min_width=12, max_width=48)
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 12
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[header_row].height = 26

    return workbook


def format_response_value(value: object) -> object:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, list):
        return ", ".join(str(part) for part in value)
    if isinstance(value, dict):
        return ", ".join(f"{key}: {part}" for key, part in value.items())
    return value


def autosize_columns(sheet, *, min_width: int, max_width: int) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            cell_length = max(len(line) for line in str(cell.value).splitlines())
            length = max(length, cell_length)
        sheet.column_dimensions[column_letter].width = min(max(length + 2, min_width), max_width)
