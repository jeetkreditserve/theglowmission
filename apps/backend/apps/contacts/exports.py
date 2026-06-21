from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone

from apps.contacts.models import Contact


BRAND_FILL = PatternFill("solid", fgColor="2B2623")
HEADER_FILL = PatternFill("solid", fgColor="D9B88C")
SUBTLE_FILL = PatternFill("solid", fgColor="F6EEE4")
WHITE_FONT = Font(color="FFF7F0", bold=True)
TITLE_FONT = Font(color="FFF7F0", bold=True, size=16)
HEADER_FONT = Font(color="2B2623", bold=True)
BODY_FONT = Font(color="2B2623")
THIN_BOTTOM = Border(bottom=Side(style="thin", color="D9B88C"))


def build_contacts_workbook(contacts) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contacts"
    sheet.sheet_view.showGridLines = False

    contacts = list(contacts)
    headers = [
        "Contact ID",
        "Status",
        "Full name",
        "Email",
        "Phone",
        "Address",
        "Age",
        "Skin type",
        "Preferred ritual",
        "Preferred day",
        "Skin goal",
        "Marketing consent",
        "Source responses",
        "First activity",
        "Last activity",
    ]
    last_column = get_column_letter(len(headers))

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"] = "The Glow Mission"
    sheet["A1"].fill = BRAND_FILL
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].alignment = Alignment(horizontal="center")
    for column in range(2, len(headers) + 1):
        sheet.cell(row=1, column=column).fill = BRAND_FILL

    sheet["A2"] = "Contacts export"
    sheet["A3"] = f"Exported: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')}"
    sheet["A4"] = f"Contact count: {len(contacts)}"
    for row in range(2, 5):
        sheet.cell(row=row, column=1).font = HEADER_FONT
        sheet.cell(row=row, column=1).fill = SUBTLE_FILL
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))

    header_row = 6
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BOTTOM

    for row_index, contact in enumerate(contacts, start=header_row + 1):
        row_values = [
            contact.id,
            contact.status.name if contact.status else "",
            contact.full_name,
            contact.email,
            contact.phone,
            contact.address,
            contact.age,
            contact.skin_type,
            contact.preferred_ritual,
            contact.preferred_day,
            contact.skin_goal,
            "Yes" if contact.marketing_consent else "No",
            contact.source_response_count,
            local_naive(contact.first_activity_at),
            local_naive(contact.last_activity_at),
        ]
        for column, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in {14, 15}:
                cell.number_format = "yyyy-mm-dd hh:mm"

    last_row = max(header_row + 1, header_row + len(contacts))
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
    sheet.freeze_panes = f"C{header_row + 1}"
    autosize_columns(sheet, min_width=12, max_width=48)
    sheet.column_dimensions["A"].width = 12
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[header_row].height = 26
    return workbook


def local_naive(value):
    if not value:
        return None
    return timezone.localtime(value).replace(tzinfo=None)


def autosize_columns(sheet, *, min_width: int, max_width: int) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            cell_length = max(len(line) for line in (str(cell.value).splitlines() or [""]))
            length = max(length, cell_length)
        sheet.column_dimensions[column_letter].width = min(max(length + 2, min_width), max_width)
